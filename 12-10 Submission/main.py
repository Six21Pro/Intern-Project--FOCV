#This code expands upon code from this source: https://learndataanalysis.org/source-code-search-nearby-businesses-with-google-maps-api-and-python/#google_vignette#

import os#this for secrets in replit
import time
import googlemaps  # pip install googlemaps
import pandas as pd  # pip install pandas
import openpyxl #for excel manipulation
from contextCreator import Contexter


def miles_to_meters(miles):
    try:
        return miles * 1_609.344
    except:
        return 0

API_KEY = os.environ['APIKEY']#this uses a secret in replit
map_client = googlemaps.Client(API_KEY)

# Address to search nearby
address = '2138 Croasdale Rd, Stroudsburg, PA'
geocode = map_client.geocode(address=address)
(lat, lng) = map(geocode[0]['geometry']['location'].get, ('lat', 'lng'))

# Search parameters
#search_string = 'pizza' #for initial test
search_string = input("input search string ")
howFar = input("how far in miles? ")

#distance = miles_to_meters(1)
distance = miles_to_meters(float(howFar))

business_list = []

# Initial request to get places
response = map_client.places_nearby(location=(lat, lng),
                                    keyword=search_string,
                                    radius=distance)

business_list.extend(response.get('results'))
next_page_token = response.get('next_page_token')

# Pagination to get all results
while next_page_token:
    time.sleep(3)  # Delay to avoid hitting rate limits#from original code
    response = map_client.places_nearby(location=(lat, lng),
                                        keyword=search_string,
                                        radius=distance,
                                        page_token=next_page_token)
    business_list.extend(response.get('results'))
    next_page_token = response.get('next_page_token')

# Create a DataFrame from the business list
df = pd.DataFrame(business_list)

# Retrieve website URLs for each place
def get_website_url(place_id):
    try:
        place_details = map_client.place(place_id)
        website_url = place_details['result'].get('website',
                                                  'No website available')
        return website_url
    except Exception as e:
        return 'Error retrieving website'

# this will get website URLs for each business
df['website'] = df['place_id'].apply(get_website_url)

# Generate Google Maps URLs#from example code
df['url'] = 'https://www.google.com/maps/place/?q=place_id:' + df['place_id']

print(f"Saved {len(df)} businesses to Excel.")

df.to_excel('temp.xlsx', index=False)#saves businesses to excel

df = pd.read_excel('temp.xlsx')  

urlsToScrape = df['website'].tolist()#put the website column in a list
print(urlsToScrape)

contextAdder = Contexter()
contextArray = []
contextArray = contextAdder.addContext2(urlsToScrape)
print(contextArray)
df['Charitable'] = contextArray#add a column named Charitable, which contains context for the websites

df.to_excel('temp.xlsx', index=False)#resaves businesses to excel

try:#if the file exists, it load it
    giving_wb = openpyxl.load_workbook('giving.xlsx')
except FileNotFoundError:
    print("File not found, creating a new one")
    giving_wb = openpyxl.Workbook()
    giving_wb.save('giving.xlsx')
    giving_wb = openpyxl.load_workbook('giving.xlsx')

# Load source and workbooks
source_wb = openpyxl.load_workbook('temp.xlsx')
giving_wb = openpyxl.load_workbook('giving.xlsx')

source_sheet = source_wb['Sheet1']  #get the sheet that was created in temp.xlsx

giving_sheet = giving_wb.create_sheet(search_string)  # New sheet name based on search_string

# Copy content from the source sheet to the new giving sheet
for row in source_sheet.iter_rows():
    for cell in row:
        giving_sheet[cell.coordinate].value = cell.value# Copy cell value

giving_wb.save('giving.xlsx')# Save the giving workbook with the copied sheet